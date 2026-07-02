import io
from datetime import datetime, timedelta
from telegram import Update
from telegram.ext import ContextTypes
from database import get_conn
from utils import require_rol

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def parsear_fecha(texto):
    try:
        return datetime.strptime(texto, "%d/%m/%Y")
    except ValueError:
        return None

def lunes_de(fecha: datetime) -> datetime:
    return fecha - timedelta(days=fecha.weekday())

def build_resumen_dia(fecha: datetime, conn):
    cur = conn.cursor()
    fecha_str = fecha.date()

    cur.execute("""
        SELECT COALESCE(SUM(monto_total), 0), COALESCE(SUM(monto_pendiente), 0), COUNT(*)
        FROM servicios
        WHERE DATE(created_at) = %s AND is_active = TRUE AND estado != 'anulado'
    """, (fecha_str,))
    total, pendiente, count = cur.fetchone()
    cobrado = float(total) - float(pendiente)

    cur.execute("""
        SELECT u.nombre, COUNT(*), COALESCE(SUM(s.monto_total - s.monto_pendiente), 0)
        FROM servicios s JOIN usuarios u ON s.mecanico_id = u.id
        WHERE DATE(s.created_at) = %s AND s.is_active = TRUE AND s.estado != 'anulado'
        GROUP BY u.nombre ORDER BY 3 DESC
    """, (fecha_str,))
    por_mecanico = cur.fetchall()

    cur.execute("""
        SELECT tipo, COALESCE(SUM(monto), 0) FROM gastos
        WHERE DATE(created_at) = %s AND is_active = TRUE GROUP BY tipo
    """, (fecha_str,))
    gastos_rows = dict(cur.fetchall())
    total_gastos = float(gastos_rows.get("gasto", 0))
    total_adelantos = float(gastos_rows.get("adelanto", 0))
    neto = cobrado - total_gastos - total_adelantos

    cur.execute("""
        SELECT s.tricimoto_num, s.tricimoto_compania, s.monto_total, s.monto_pendiente, s.descripcion, u.nombre
        FROM servicios s JOIN usuarios u ON s.mecanico_id = u.id
        WHERE DATE(s.created_at) = %s AND s.is_active = TRUE AND s.estado != 'anulado'
        ORDER BY s.created_at ASC
    """, (fecha_str,))
    detalle = cur.fetchall()
    cur.close()

    return {
        "total": float(total), "cobrado": cobrado, "pendiente": float(pendiente),
        "count": int(count), "gastos": total_gastos, "adelantos": total_adelantos,
        "neto": neto, "por_mecanico": por_mecanico, "detalle": detalle
    }

def build_resumen_semana(lunes: datetime, conn):
    cur = conn.cursor()
    domingo = lunes + timedelta(days=6)

    cur.execute("""
        SELECT COALESCE(SUM(monto_total), 0), COALESCE(SUM(monto_pendiente), 0), COUNT(*)
        FROM servicios
        WHERE DATE(created_at) BETWEEN %s AND %s AND is_active = TRUE AND estado != 'anulado'
    """, (lunes.date(), domingo.date()))
    total, pendiente, count = cur.fetchone()
    cobrado = float(total) - float(pendiente)

    cur.execute("""
        SELECT u.nombre, COUNT(*), COALESCE(SUM(s.monto_total - s.monto_pendiente), 0),
               COALESCE((SELECT SUM(g.monto) FROM gastos g WHERE g.registrado_por = u.id
                         AND g.tipo = 'adelanto' AND DATE(g.created_at) BETWEEN %s AND %s
                         AND g.is_active = TRUE), 0)
        FROM servicios s JOIN usuarios u ON s.mecanico_id = u.id
        WHERE DATE(s.created_at) BETWEEN %s AND %s AND s.is_active = TRUE AND s.estado != 'anulado'
        GROUP BY u.id, u.nombre ORDER BY 3 DESC
    """, (lunes.date(), domingo.date(), lunes.date(), domingo.date()))
    por_mecanico = cur.fetchall()

    cur.execute("""
        SELECT tipo, COALESCE(SUM(monto), 0) FROM gastos
        WHERE DATE(created_at) BETWEEN %s AND %s AND is_active = TRUE GROUP BY tipo
    """, (lunes.date(), domingo.date()))
    gastos_rows = dict(cur.fetchall())
    total_gastos = float(gastos_rows.get("gasto", 0))
    total_adelantos = float(gastos_rows.get("adelanto", 0))
    neto = cobrado - total_gastos - total_adelantos

    cur.close()
    return {
        "lunes": lunes, "domingo": domingo,
        "total": float(total), "cobrado": cobrado, "pendiente": float(pendiente),
        "count": int(count), "gastos": total_gastos, "adelantos": total_adelantos,
        "neto": neto, "por_mecanico": por_mecanico
    }

def fmt_resumen_dia(data, fecha):
    titulo = f"📊 *Resumen del {fecha.strftime('%d/%m/%Y')}*\n\n"
    msg = (f"🔧 Servicios: {data['count']}\n"
           f"💰 Total: ${data['total']:.2f}\n"
           f"✅ Cobrado: ${data['cobrado']:.2f}\n"
           f"⏳ Pendiente: ${data['pendiente']:.2f}\n"
           f"🛒 Gastos: ${data['gastos']:.2f}\n"
           f"💵 Adelantos: ${data['adelantos']:.2f}\n"
           f"📦 Neto: ${data['neto']:.2f}\n")
    if data["por_mecanico"]:
        msg += "\n🔧 *Por mecánico:*\n"
        for nombre, count, cobrado in data["por_mecanico"]:
            msg += f"  • {nombre}: {count} servicios — ${float(cobrado):.2f}\n"
    if data["detalle"]:
        msg += "\n📋 *Detalle:*\n"
        for num, color, total, pend, desc, mec in data["detalle"]:
            estado_emoji = "✅" if float(pend) == 0 else "⏳"
            msg += f"{estado_emoji} {num} {color} — ${float(total):.2f}"
            if desc: msg += f" — {desc}"
            msg += f" ({mec})\n"
    return titulo + msg

def fmt_resumen_semana(data):
    lunes = data["lunes"].strftime("%d/%m/%Y")
    domingo = data["domingo"].strftime("%d/%m/%Y")
    titulo = f"📅 *Semana del {lunes} al {domingo}*\n\n"
    msg = (f"🔧 Servicios: {data['count']}\n"
           f"💰 Total: ${data['total']:.2f}\n"
           f"✅ Cobrado: ${data['cobrado']:.2f}\n"
           f"⏳ Pendiente: ${data['pendiente']:.2f}\n"
           f"🛒 Gastos: ${data['gastos']:.2f}\n"
           f"💵 Adelantos: ${data['adelantos']:.2f}\n"
           f"📦 Neto: ${data['neto']:.2f}\n")
    if data["por_mecanico"]:
        msg += "\n🔧 *Por mecánico:*\n"
        for nombre, count, cobrado, adelanto in data["por_mecanico"]:
            msg += f"  • {nombre}: {count} servicios — ${float(cobrado):.2f} (adelanto: ${float(adelanto):.2f})\n"
    return titulo + msg

# ─── COMANDOS DE CONSULTA (todos los roles) ───────────────────────────────────

@require_rol("admin", "jefe", "mecanico")
async def cmd_hoy(update: Update, context: ContextTypes.DEFAULT_TYPE):
    conn = get_conn()
    data = build_resumen_dia(datetime.now(), conn)
    conn.close()
    await update.message.reply_text(fmt_resumen_dia(data, datetime.now()), parse_mode="Markdown")


@require_rol("admin", "jefe", "mecanico")
async def cmd_resumen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if not args:
        await update.message.reply_text("Uso: `/resumen dd/mm/yyyy`", parse_mode="Markdown")
        return
    fecha = parsear_fecha(args[0])
    if not fecha:
        await update.message.reply_text("❌ Formato inválido. Ej: `20/06/2025`", parse_mode="Markdown")
        return
    conn = get_conn()
    data = build_resumen_dia(fecha, conn)
    conn.close()
    await update.message.reply_text(fmt_resumen_dia(data, fecha), parse_mode="Markdown")


@require_rol("admin", "jefe", "mecanico")
async def cmd_resumen_semana(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if not args:
        await update.message.reply_text("Uso: `/resumen_semana dd/mm/yyyy`", parse_mode="Markdown")
        return
    fecha = parsear_fecha(args[0])
    if not fecha:
        await update.message.reply_text("❌ Formato inválido. Ej: `20/06/2025`", parse_mode="Markdown")
        return
    lunes = lunes_de(fecha)
    conn = get_conn()
    data = build_resumen_semana(lunes, conn)
    conn.close()
    await update.message.reply_text(fmt_resumen_semana(data), parse_mode="Markdown")

# ─── EXPORTAR EXCEL (admin/jefe) ─────────────────────────────────────────────

@require_rol("admin", "jefe")
async def cmd_exportar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        await update.message.reply_text("❌ openpyxl no instalado.")
        return

    args = context.args
    if not args:
        await update.message.reply_text(
            "Uso:\n`/exportar dia dd/mm/yyyy`\n`/exportar semana dd/mm/yyyy`\n`/exportar mes mm/yyyy`",
            parse_mode="Markdown"
        )
        return

    tipo = args[0].lower()
    conn = get_conn()
    cur = conn.cursor()

    # Determinar rango de fechas
    if tipo == "dia":
        if len(args) < 2:
            await update.message.reply_text("Uso: `/exportar dia dd/mm/yyyy`", parse_mode="Markdown")
            cur.close(); conn.close(); return
        fecha = parsear_fecha(args[1])
        if not fecha:
            await update.message.reply_text("❌ Fecha inválida"); cur.close(); conn.close(); return
        fecha_ini = fecha_fin = fecha.date()
        nombre_archivo = f"reporte_{fecha.strftime('%d-%m-%Y')}.xlsx"
        titulo = f"Reporte del {fecha.strftime('%d/%m/%Y')}"

    elif tipo == "semana":
        if len(args) < 2:
            await update.message.reply_text("Uso: `/exportar semana dd/mm/yyyy`", parse_mode="Markdown")
            cur.close(); conn.close(); return
        fecha = parsear_fecha(args[1])
        if not fecha:
            await update.message.reply_text("❌ Fecha inválida"); cur.close(); conn.close(); return
        lunes = lunes_de(fecha)
        domingo = lunes + timedelta(days=6)
        fecha_ini, fecha_fin = lunes.date(), domingo.date()
        nombre_archivo = f"semana_{lunes.strftime('%d-%m-%Y')}.xlsx"
        titulo = f"Semana {lunes.strftime('%d/%m/%Y')} — {domingo.strftime('%d/%m/%Y')}"

    elif tipo == "mes":
        if len(args) < 2:
            await update.message.reply_text("Uso: `/exportar mes mm/yyyy`", parse_mode="Markdown")
            cur.close(); conn.close(); return
        try:
            fecha = datetime.strptime(args[1], "%m/%Y")
        except ValueError:
            await update.message.reply_text("❌ Formato inválido. Ej: `06/2025`", parse_mode="Markdown")
            cur.close(); conn.close(); return
        import calendar
        ultimo_dia = calendar.monthrange(fecha.year, fecha.month)[1]
        fecha_ini = fecha.date()
        fecha_fin = fecha.replace(day=ultimo_dia).date()
        nombre_archivo = f"mes_{fecha.strftime('%m-%Y')}.xlsx"
        titulo = f"Mes {fecha.strftime('%B %Y')}"

    else:
        await update.message.reply_text("❌ Tipo inválido. Usa: dia, semana, mes")
        cur.close(); conn.close(); return

    # Consultas
    cur.execute("""
        SELECT s.id, s.created_at, s.tricimoto_num, s.tricimoto_compania,
               s.monto_total, s.monto_pendiente, (s.monto_total - s.monto_pendiente) AS cobrado,
               s.descripcion, s.estado, u_mec.nombre AS mecanico, u_reg.nombre AS registrado_por
        FROM servicios s
        JOIN usuarios u_mec ON s.mecanico_id = u_mec.id
        JOIN usuarios u_reg ON s.registrado_por = u_reg.id
        WHERE DATE(s.created_at) BETWEEN %s AND %s AND s.is_active = TRUE AND s.estado != 'anulado'
        ORDER BY s.created_at ASC
    """, (fecha_ini, fecha_fin))
    servicios = cur.fetchall()

    cur.execute("""
        SELECT p.id, p.created_at, s.tricimoto_num, s.tricimoto_compania, p.monto, u.nombre AS registrado_por
        FROM pagos p
        JOIN servicios s ON p.servicio_id = s.id
        JOIN usuarios u ON p.registrado_por = u.id
        WHERE DATE(p.created_at) BETWEEN %s AND %s AND p.is_active = TRUE
        ORDER BY p.created_at ASC
    """, (fecha_ini, fecha_fin))
    pagos = cur.fetchall()

    cur.execute("""
        SELECT g.id, g.created_at, g.tipo, g.monto, g.descripcion, u.nombre AS registrado_por
        FROM gastos g JOIN usuarios u ON g.registrado_por = u.id
        WHERE g.tipo = 'gasto' AND DATE(g.created_at) BETWEEN %s AND %s AND g.is_active = TRUE
        ORDER BY g.created_at ASC
    """, (fecha_ini, fecha_fin))
    gastos = cur.fetchall()

    cur.execute("""
        SELECT g.id, g.created_at, g.descripcion AS destinatario, g.monto, u.nombre AS registrado_por
        FROM gastos g JOIN usuarios u ON g.registrado_por = u.id
        WHERE g.tipo = 'adelanto' AND DATE(g.created_at) BETWEEN %s AND %s AND g.is_active = TRUE
        ORDER BY g.created_at ASC
    """, (fecha_ini, fecha_fin))
    adelantos = cur.fetchall()

    cur.close()
    conn.close()

    # ── Construir Excel ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="111827")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
    BORDER = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    PAID_FILL  = PatternFill("solid", fgColor="F0FDF4")
    PEND_FILL  = PatternFill("solid", fgColor="FEFCE8")

    def write_sheet(ws, headers, rows, col_widths, formatters=None):
        # Título
        ws.merge_cells(f"A1:{chr(64+len(headers))}1")
        ws["A1"] = titulo
        ws["A1"].font = Font(bold=True, size=12)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Headers
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = BORDER

        # Datos
        for row_idx, row in enumerate(rows, 3):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")
                if formatters and col_idx in formatters:
                    formatters[col_idx](cell, row)

        # Anchos
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64+col_idx)].width = width

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20

    # Hoja 1: Servicios
    ws1 = wb.active
    ws1.title = "Servicios"
    headers_s = ["ID", "Fecha", "Tricimoto", "Compania", "Total", "Cobrado", "Pendiente", "Descripción", "Estado", "Mecánico", "Registrado por"]
    rows_s = [(s[0], s[1].strftime("%d/%m/%Y %H:%M"), s[2], s[3],
               float(s[4]), float(s[6]), float(s[5]), s[7] or "", s[8], s[9], s[10])
              for s in servicios]
    def fmt_servicio(cell, row):
        if cell.column == 5 or cell.column == 6 or cell.column == 7:
            cell.number_format = '"$"#,##0.00'
        if cell.column == 9:
            cell.fill = PAID_FILL if row[8] == "pagado" else PEND_FILL
    write_sheet(ws1, headers_s, rows_s, [6, 16, 10, 10, 10, 10, 10, 30, 10, 15, 15],
                {5: fmt_servicio, 6: fmt_servicio, 7: fmt_servicio, 9: fmt_servicio})

    # Hoja 2: Pagos
    ws2 = wb.create_sheet("Pagos")
    headers_p = ["ID", "Fecha", "Tricimoto", "Color", "Monto", "Registrado por"]
    rows_p = [(p[0], p[1].strftime("%d/%m/%Y %H:%M"), p[2], p[3], float(p[4]), p[5]) for p in pagos]
    def fmt_pago(cell, row):
        if cell.column == 5: cell.number_format = '"$"#,##0.00'
    write_sheet(ws2, headers_p, rows_p, [6, 16, 10, 10, 12, 15], {5: fmt_pago})

    # Hoja 3: Gastos
    ws3 = wb.create_sheet("Gastos")
    headers_g = ["ID", "Fecha", "Tipo", "Monto", "Descripción", "Registrado por"]
    rows_g = [(g[0], g[1].strftime("%d/%m/%Y %H:%M"), g[2], float(g[3]), g[4] or "", g[5]) for g in gastos]
    def fmt_gasto(cell, row):
        if cell.column == 4: cell.number_format = '"$"#,##0.00'
    write_sheet(ws3, headers_g, rows_g, [6, 16, 10, 12, 30, 15], {4: fmt_gasto})

    # Hoja 4: Adelantos
    ws4 = wb.create_sheet("Adelantos")
    headers_a = ["ID", "Fecha", "Destinatario", "Monto", "Registrado por"]
    rows_a = [(a[0], a[1].strftime("%d/%m/%Y %H:%M"), a[2] or "", float(a[3]), a[4]) for a in adelantos]
    def fmt_adelanto(cell, row):
        if cell.column == 4: cell.number_format = '"$"#,##0.00'
    write_sheet(ws4, headers_a, rows_a, [6, 16, 20, 12, 15], {4: fmt_adelanto})

    # Hoja 5: Resumen
    ws5 = wb.create_sheet("Resumen")
    ws5["A1"] = titulo
    ws5["A1"].font = Font(bold=True, size=13)
    ws5.merge_cells("A1:B1")

    total_servicios = sum(float(s[4]) for s in servicios)
    total_cobrado   = sum(float(s[6]) for s in servicios)
    total_pendiente = sum(float(s[5]) for s in servicios)
    total_gastos    = sum(float(g[3]) for g in gastos)
    total_adelantos = sum(float(a[3]) for a in adelantos)
    neto = total_cobrado - total_gastos - total_adelantos

    resumen_rows = [
        ("Total servicios registrados", len(servicios)),
        ("Monto total servicios", total_servicios),
        ("Total cobrado", total_cobrado),
        ("Total pendiente", total_pendiente),
        ("Total gastos", total_gastos),
        ("Total adelantos", total_adelantos),
        ("Neto", neto),
    ]
    for i, (label, val) in enumerate(resumen_rows, 3):
        ws5[f"A{i}"] = label
        ws5[f"A{i}"].font = Font(bold=True)
        ws5[f"B{i}"] = val
        if i > 3:
            ws5[f"B{i}"].number_format = '"$"#,##0.00'
        ws5[f"A{i}"].border = BORDER
        ws5[f"B{i}"].border = BORDER

    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 18

    # Enviar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await update.message.reply_document(
        document=buffer,
        filename=nombre_archivo,
        caption=f"📊 {titulo}\n\n🔧 {len(servicios)} servicios | ✅ ${total_cobrado:.2f} cobrado | 📦 Neto: ${neto:.2f}"
    )
